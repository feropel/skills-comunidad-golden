# -*- coding: utf-8 -*-
"""Califica pedidos Dropi: direccion + cliente + transportadora (precio/efectividad)."""
import json, csv, re, unicodedata, collections, sys
import os
try:
    import openpyxl
except ModuleNotFoundError:
    sys.exit('Falta openpyxl (lee los .xlsx de Dropi). Instalar una vez:  pip3 install openpyxl')
DATA = os.environ.get("DROPI_DATA") or os.path.expanduser("~/DROPI-LOGISTICA")
if not os.path.isdir(os.path.join(DATA, 'datos')):
    sys.exit('No encuentro el cerebro de datos en %s.\n'
             'Apunta DROPI_DATA a tu carpeta DROPI-LOGISTICA:\n'
             '  export DROPI_DATA="/ruta/a/DROPI-LOGISTICA"' % DATA)
def D(p): return os.path.join(DATA, p)
def huellas_recientes():
    import glob
    f=sorted(glob.glob(D('datos/huellas/*.psv')))
    if not f: raise SystemExit('No hay huellas capturadas en datos/huellas/. Corre primero la cosecha del panel.')
    return f[-1]


def N(s):
    s=unicodedata.normalize('NFD',str(s) if s is not None else '')
    return ''.join(c for c in s if unicodedata.category(c)!='Mn').upper().strip()
FL={(N(r['ciu']),N(r['dep'])):r for r in json.load(open(D('datos/FLETES-CALI.json')))['R']}
EF={N(k):v for k,v in json.load(open(D('datos/EFECTIVIDAD-PLATAFORMA.json'))).items()}
FACT={1:0.966,2:1.068,3:1.268,4:1.290}
# costo real del retorno: fraccion medida por transportadora (datos/COSTO-RETORNO.json).
# Trampa 5: solo se usa la fraccion con confianza alta o media; el resto asume 1.00 (conservador).
# Un caso suelto en $0 NO es "no cobra retorno": ya produjo cuatro recomendaciones equivocadas.
RET_DEF=1.00
try:
    _cr=json.load(open(D('datos/COSTO-RETORNO.json')))
    _med=_cr[sorted(k for k in _cr if k.startswith('medido'))[-1]]
    RET={N(k):v['fraccion'] for k,v in _med.items() if str(v.get('confianza','')).lower() in ('alta','media')}
except Exception:
    RET={'INTERRAPIDISIMO':1.00,'ENVIA':0.72}   # ultimo medido conocido, por si falta el archivo
# rechazos reales de la empresa de fulfillment: mandan sobre la cotizacion de Dropi
try:
    _R=json.load(open(D('datos/RECHAZOS-FULFILLMENT.json')))['rechazos']
except Exception:
    _R=[]
VETADAS_CIUDAD={(N(r['transportadora']),N(r['ciudad']),N(r['departamento'])) for r in _R if r['alcance']=='CIUDAD'}
VETADAS_DEPTO={(N(r['transportadora']),N(r['departamento'])) for r in _R if r['alcance']=='DEPARTAMENTO'}
VETADAS_PAIS={N(r['transportadora']) for r in _R if r['alcance']=='NACIONAL'}
def vetada(car,ciu,dep):
    return (car,ciu,dep) in VETADAS_CIUDAD or (car,dep) in VETADAS_DEPTO or car in VETADAS_PAIS
OFICINA_OK={'INTERRAPIDISIMO','COORDINADORA'}          # colombia.md: solo estas dos
HABILITADAS={'INTERRAPIDISIMO','ENVIA','TCC','VELOCES','COORDINADORA','DOMINA','JAMV-DRIVE'}
# GD1.1: la bodega manda — TRANSPORTADORAS-OPERATIVAS.json se antepone a todo calculo.
# NUNCA USADA no se recomienda aunque gane el calculo; MARGINAL se propone avisando el riesgo.
try:
    _op=json.load(open(D('datos/TRANSPORTADORAS-OPERATIVAS.json')))
    _opm=_op[sorted(k for k in _op if k.startswith('medido'))[-1]]
    ESTADO_BODEGA={N(k):v['estado'] for k,v in _opm.items() if isinstance(v,dict) and 'estado' in v}
except Exception:
    ESTADO_BODEGA={}
    print('AVISO: sin TRANSPORTADORAS-OPERATIVAS.json — no se puede filtrar por bodega; verificar a mano', file=sys.stderr)
def bodega_ok(car): return (not ESTADO_BODEGA) or ESTADO_BODEGA.get(car) in ('OPERATIVA','MARGINAL')

def analiza_dir(d, ciudad):
    """Devuelve (estado, detalle, transportadora_forzada)"""
    t=N(d); orig=str(d or '')
    # 1. recogida en oficina
    if 'OFICINA' in t or 'RETIRA EN' in t or 'RECOGE EN' in t:
        comp = t.replace(' ','').replace('.','')
        if 'INTERRAPID' in comp or 'INTERRAPIS' in comp:
            return ('OFICINA','retiro en oficina InterRapidisimo — la transportadora NO se puede cambiar','INTERRAPIDISIMO')
        if 'COORDINADORA' in comp:
            return ('OFICINA','retiro en oficina Coordinadora — la transportadora NO se puede cambiar','COORDINADORA')
        if 'SERVIENTREGA' in comp:
            return ('BLOQUEO','pide oficina Servientrega, que NO esta habilitada — hay que reconfirmar con el cliente',None)
        return ('BLOQUEO','dice oficina pero no se identifica cual — confirmar transportadora y ciudad',None)
    # 2. transportadora ajena mencionada en la direccion
    for x in ['SERVIENTREGA','RAPIENTREGA','DEPRISA','SATURNO']:
        if x in t: return ('REVISAR','la direccion menciona "%s"; si es punto de esa empresa no sirve — confirmar'%x.title(),None)
    # 3. nomenclatura
    # numero de puerta completo: "# 30 - 45", "#30A-45", "# 1 OESTE - 03"
    puerta = bool(re.search(r'#\s*\d+\s*[A-Z]?\s*(OESTE|ESTE|SUR|NORTE|BIS)?\s*-\s*\d+', t)) or bool(re.search(r'\b\d+\s*[A-Z]?\s*-\s*\d+', t))
    mzcs = bool(re.search(r'\b(MZ|MANZANA)\b.*\b(CS|CASA)\b', t))
    tiene_num = puerta or mzcs
    # dos direcciones en el mismo campo
    if len(re.findall(r'#', t))>=2:
        return ('REVISAR','la direccion trae DOS nomenclaturas distintas — confirmar cual es la buena',None)
    if re.search(r'#\s*\d+\s*[A-Z]?\s*$', t) or (re.search(r'#\s*\d', t) and not puerta):
        return ('INCOMPLETA','la nomenclatura queda cortada: falta el numero de puerta despues del guion',None)
    tiene_via = bool(re.search(r'\b(CL|CLL|CALLE|KR|CRA|CR|CARRERA|AV|AVENIDA|DG|DIAGONAL|TV|TRANSVERSAL|AK|AC|MZ|MANZANA|VEREDA|KM)\b', t))
    tres = len(re.findall(r'\d+', t))>=3
    if not tiene_via and not tiene_num:
        return ('INCOMPLETA','no hay nomenclatura: falta via y numeros (Calle/Carrera # __-__)',None)
    if tiene_via and not tiene_num and not tres:
        return ('INCOMPLETA','falta el numero de puerta: solo trae la via',None)
    if not tiene_num and tres:
        return ('REVISAR','numeracion sin formato # __-__; verificar que sea la puerta correcta',None)
    # 4. vivienda colectiva sin numero interno
    colect=re.search(r'\b(CONJUNTO|CONJ|UNIDAD|EDIFICIO|EDIFICO|ED|TORRE|TO|BLOQUE|BL)\b', t)
    interno=re.search(r'\b(AP|APTO|APTM|APARTAMENTO|CASA|CS|INT|INTERIOR|PISO|PI|LOCAL|OF)\b\s*\.?\s*\d*', t)
    if colect and not interno:
        return ('INCOMPLETA','menciona %s pero no da torre/apartamento'%colect.group(1).lower(),None)
    # 5. lugar publico / comercial sin punto
    if re.search(r'\b(ESTACION DE POLICIA|HOTEL|TERMINAL|CENTRO DE NEGOCIOS|CC|CENTRO COMERCIAL)\b', t) and not interno:
        return ('REVISAR','entrega en lugar publico o comercial sin oficina/local — confirmar a quien se entrega',None)
    return ('OK','',None)

def rutas_export():
    """Las 2 rutas de los export de Dropi. Por argumento, por variable de entorno, o
    el más reciente que encuentre. NUNCA una ruta fija: esto corre en la máquina de
    cualquier alumna, no solo en la del dueño."""
    import glob
    a = sys.argv[1] if len(sys.argv)>1 else os.environ.get('DROPI_ORDENES')
    b = sys.argv[2] if len(sys.argv)>2 else os.environ.get('DROPI_PRODUCTOS')
    if not a:
        c = sorted(glob.glob(os.path.expanduser('~/Desktop/ordenes_*.xlsx')))
        c = [x for x in c if 'productos' not in os.path.basename(x)]
        a = c[-1] if c else None
    if not b:
        c = sorted(glob.glob(os.path.expanduser('~/Desktop/ordenes_productos_*.xlsx')))
        b = c[-1] if c else None
    if not a or not b:
        sys.exit("Faltan los export de Dropi. Uso:\n"
                 "  python3 calificar.py <ordenes.xlsx> <ordenes_productos.xlsx>\n"
                 "o exporta DROPI_ORDENES y DROPI_PRODUCTOS.\n"
                 "Se descargan de Dropi > Ordenes > Exportar (los DOS archivos).")
    return a, b

def cargar():
    A, B = rutas_export()
    wb=openpyxl.load_workbook(A,read_only=True,data_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True)); H={c:i for i,c in enumerate(rows[0])}
    d=[r for r in rows[1:] if r[H['ESTATUS']] in ('PENDIENTE','PENDIENTE CONFIRMACION')]; wb.close()
    wb2=openpyxl.load_workbook(B,read_only=True,data_only=True); ws2=wb2.active
    r2=list(ws2.iter_rows(values_only=True)); H2={c:i for i,c in enumerate(r2[0])}
    q=collections.Counter()
    for r in r2[1:]:
        if r[H2['ESTATUS']] in ('PENDIENTE','PENDIENTE CONFIRMACION'): q[str(r[H2['ID']])]+=int(r[H2['CANTIDAD']] or 0)
    wb2.close(); return d,H,q

data,H,QTY=cargar()
hue={m['orden']:m for m in csv.DictReader(open(huellas_recientes()),delimiter='|')}
out=[]
for r in data:
    oid=str(r[H['ID']]); ciu=N(r[H['CIUDAD DESTINO']]); dep=N(r[H['DEPARTAMENTO DESTINO']]); asig=N(r[H['TRANSPORTADORA']])
    prepago = r[H['TIPO DE ENVIO']]=='SIN RECAUDO'
    ticket=float(r[H['VALOR DE COMPRA EN PRODUCTOS']] or 0); costo=float(r[H['TOTAL EN PRECIOS DE PROVEEDOR']] or 0)
    freal=float(r[H['PRECIO FLETE']] or 0); q=QTY.get(oid,1); f=FACT.get(q,1.29)
    est,det,forzada=analiza_dir(r[H['DIRECCION']], ciu)
    h=hue.get(oid,{}); cli={}
    for tok in (h.get('por_transportadora') or '').split(';'):
        if ':' in tok:
            k,v=tok.split(':'); e,d=v.split('/'); cli[N(k)]=(int(e),int(d))
    tot=int(h.get('total') or 0); entP=float(h.get('entP') or 0); devP=float(h.get('devP') or 0)
    fl=FL.get((ciu,dep)); ef=EF.get(dep,{})
    cands=[]
    for car,precio in (fl['m'].items() if fl else []):
        if car not in HABILITADAS: continue
        if not bodega_ok(car): continue
        base=ef.get(car,{}).get('pct')
        if base is None: continue
        if forzada and car!=forzada: continue
        if vetada(car,ciu,dep): continue
        p=base/100
        e,d=cli.get(car,(0,0))
        if not prepago:
            # 1) la zona se ajusta con el historial GLOBAL del cliente (la zona pesa como 10 envios)
            if tot>0: p=(entP/100*tot + 10*p)/(tot+10)
            # 2) y luego con lo que ese cliente vivio CON ESA transportadora (la zona pesa como 3)
            n=e+d
            if n>0: p=(e + 3*p)/(n+3)
        else:
            p=max(p,0.97)                      # 9 de 9 prepagos resueltos se entregaron
        flete = freal if car==asig else round(precio*f)
        retorno = 0 if prepago else flete*RET.get(car,RET_DEF)
        margen=ticket-costo-flete
        ev=p*margen-(1-p)*(flete+retorno)
        cands.append(dict(car=car,ev=round(ev),flete=flete,base=round(base,2),p=round(p*100,1),ret=round(retorno),
                          marg=ESTADO_BODEGA.get(car)=='MARGINAL'))
    if prepago:   # prioridad precio, con piso de efectividad
        if cands:
            mejorEf=max(c['base'] for c in cands)
            viables=[c for c in cands if c['base']>=mejorEf-3]
            viables.sort(key=lambda c:c['flete'])
            barato=viables[0]
            # si otra con mejor efectividad cuesta menos de $1.500 mas, se prefiere
            up=[c for c in viables if c['base']>barato['base'] and c['flete']-barato['flete']<=1500]
            best=max(up,key=lambda c:c['base']) if up else barato
            cands.sort(key=lambda c:c['flete'])
        else: best=None
    else:
        cands.sort(key=lambda c:-c['ev']); best=cands[0] if cands else None
    a=next((c for c in cands if c['car']==asig),None)
    out.append(dict(orden=oid,estatus=r[H['ESTATUS']],cli=r[H['NOMBRE CLIENTE']],dir=r[H['DIRECCION']],ciu=ciu,dep=dep,
        asig=asig,prepago=prepago,q=q,ticket=ticket,costo=costo,freal=freal,
        dirEstado=est,dirDetalle=det,forzada=forzada,
        tot=tot,entP=entP,devP=devP,prob=h.get('prob',''),tipo=h.get('tipo',''),
        cands=cands,best=best,asigC=a,delta=(best['ev']-a['ev']) if (best and a and not prepago) else ((a['flete']-best['flete']) if (best and a and prepago) else None)))
json.dump(out,open(D('salidas/salida-v3.json'),'w'),ensure_ascii=False,indent=1)
print('analizados:',len(out))
