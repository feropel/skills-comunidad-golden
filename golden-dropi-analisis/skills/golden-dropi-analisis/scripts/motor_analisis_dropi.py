# -*- coding: utf-8 -*-
"""
MOTOR DE ANÁLISIS DROPI — skill golden-dropi-analisis
Lee los exports crudos de Dropi (por pedido = 63 cols, por producto = 53 cols),
los agrupa por cuenta/periodo y genera 2 maestros Excel en <base>/Analisis/:
    MAESTRO_LOGISTICA.xlsx   (KPIs de entrega, transportadoras, ciudades, novedades, evolución)
    MAESTRO_CONTACTOS.xlsx   (clientes dedup por teléfono, % efectividad, segmentos, enriquecido)

USO:
    python3 motor_analisis_dropi.py "<carpeta_base>"
  <carpeta_base> = carpeta que contiene los exports. Puede tener:
    - subcarpetas por cuenta (ej. Hotmail/, Gmail/, Cliente-A/) con los .xlsx dentro, o
    - los .xlsx sueltos directamente (se tratan como una sola cuenta "General").
  Si no se pasa carpeta, usa el directorio actual.

CONFIG OPCIONAL — <base>/_config_dropi.json:
    {
      "test_phones": ["3001234567"],           # teléfonos de prueba a excluir (dueño/tester)
      "test_name_keywords": ["PRUEBA","TEST"],  # nombres a excluir
      "currency": "$"
    }

ENRIQUECIMIENTO OPCIONAL (si existen, se usan; si no, se omiten sin error) en
<base>/Analisis/_FUENTES/:
    - "Etiquetas *.csv" / "Contactos *.csv"  (export de WhatsApp con columnas phone, labels, saved_name)
    - "NO CLIENTES*.csv"                       (leads de un bot tipo Chatea/ManyChat)
    - "BASE*.xlsx" con una hoja que contenga "POSIBLE" (lista para mensaje masivo)

Requiere: openpyxl.
"""
import glob, os, re, csv, sys, json
from datetime import datetime
from collections import defaultdict, Counter
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Falta la librería 'openpyxl'. Instálala con:  pip install openpyxl")

# ------------------------------------------------------------------ config
BASE = os.path.abspath(sys.argv[1]) if len(sys.argv) > 1 else os.getcwd()
OUT  = os.path.join(BASE, "Analisis")
SRC  = os.path.join(OUT, "_FUENTES")
os.makedirs(OUT, exist_ok=True)

CFG = {"test_phones": [], "test_name_keywords": ["PRUEBA", "TEST"], "currency": "$",
       "gasto_publicidad": None, "negocio": "GOLDEN"}
_cfgp = os.path.join(BASE, "_config_dropi.json")
if os.path.exists(_cfgp):
    try: CFG.update(json.load(open(_cfgp, encoding="utf-8")))
    except Exception as e: print("Aviso: _config_dropi.json no se pudo leer:", e)
CUR = CFG.get("currency", "$")
TEST_PHONES = set(re.sub(r"\D", "", str(p))[-10:] for p in CFG.get("test_phones", []))
TEST_KW = [k.upper() for k in CFG.get("test_name_keywords", ["PRUEBA", "TEST"])]

# ------------------------------------------------------------------ helpers
def g(r, i): return r[i] if (i is not None and i < len(r)) else None
def up(x): return str(x).strip().upper() if x is not None else ""
def parse_date(v):
    if isinstance(v, datetime): return v
    if v is None: return None
    s = str(v).strip()
    for f in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try: return datetime.strptime(s[:10], f)
        except: pass
    return None
def money(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[^\d\-,\.]", "", str(v))
    if not s or s in ("-", ".", ","): return 0.0
    s = s.replace(".", "").replace(",", ".")   # formato LatAm: punto = miles
    try: return float(s)
    except: return 0.0
def phone_key(v):
    d = re.sub(r"\D", "", str(v or ""))
    if len(d) > 10 and d.startswith("57"): d = d[2:]
    return d[-10:] if len(d) >= 10 else d
def es_prueba(nombre, telefono=None):
    n = up(nombre)
    if any(k in n for k in TEST_KW): return True
    if telefono is not None and phone_key(telefono) in TEST_PHONES: return True
    return False
def classify(estatus):
    e = up(estatus)
    if e == "ENTREGADO": return "entregado"
    if "DEVOLUC" in e or "REEXPEDICION" in e: return "devolucion"
    if e in ("CANCELADO", "RECHAZADO"): return "cancelado"
    return "transito"

def pipeline(estatus, clase):
    """Estado de FLUJO/caja: dónde está la plata de esta orden hoy."""
    if clase == "entregado": return "realizado"      # cobrado
    if clase == "devolucion": return "devuelto"       # perdido (flete de ida y vuelta)
    if clase == "cancelado": return "cancelado"       # nunca salió, no cuenta
    e = up(estatus)                                    # clase == transito
    if any(w in e for w in ("DEVOLUC", "REEXPED", "RECOGIDA FALLIDA")): return "en_camino_devol"
    return "en_camino"                                 # potencial, aún sin definir

def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x is not None for x in r)]
    wb.close()
    if not rows: return [], {}
    h = [str(x).strip() if x else "" for x in rows[0]]
    return rows[1:], {c: i for i, c in enumerate(h)}

def discover(base):
    """Devuelve dos listas de (path, cuenta, headers_idx-kind). Clasifica por columnas."""
    peds, prods = [], []
    for root, dirs, fns in os.walk(base):
        b = os.path.basename(root)
        if b in ("Analisis", "_FUENTES", "__MACOSX"): dirs[:] = []; continue
        rel = os.path.relpath(root, base)
        cuenta = "General" if rel == "." else rel.split(os.sep)[0]
        for fn in sorted(fns):
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$") or fn.startswith("_"): continue
            path = os.path.join(root, fn)
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb[wb.sheetnames[0]]
                head = [str(x).strip().upper() if x else "" for x in next(ws.iter_rows(values_only=True))]
                wb.close()
            except Exception:
                continue
            if "ESTATUS" not in head or "ID" not in head:   # no es export Dropi
                continue
            (prods if "PRODUCTO" in head else peds).append((path, cuenta))
    return peds, prods

def period_label(path, fecha_min):
    m = re.search(r"(\d{4})-(\d{2})", os.path.basename(path))
    if m: return m.group(0)
    return f"{fecha_min:%Y-%m}" if fecha_min else "0000-00"

# ------------------------------------------------------------------ carga
peds_f, prods_f = discover(BASE)
if not peds_f and not prods_f:
    print("No se encontraron exports de Dropi en:", BASE); sys.exit(1)

ped, prod = [], []
for path, acc in peds_f:
    rows, idx = load_rows(path)
    fmin = min((parse_date(g(r, idx.get("FECHA"))) for r in rows if parse_date(g(r, idx.get("FECHA")))), default=None)
    bl = period_label(path, fmin)
    for r in rows:
        if es_prueba(g(r, idx.get("NOMBRE CLIENTE")), g(r, idx.get("TELÉFONO"))): continue
        est = g(r, idx.get("ESTATUS"))
        _cl = classify(est)
        ped.append(dict(cuenta=acc, bimestre=bl, oid=str(g(r, idx.get("ID"))),
            fecha=parse_date(g(r, idx.get("FECHA"))), estatus=up(est), clase=_cl,
            flujo=pipeline(est, _cl),
            trans=up(g(r, idx.get("TRANSPORTADORA"))) or "(SIN DATO)",
            depto=up(g(r, idx.get("DEPARTAMENTO DESTINO"))) or "(SIN DATO)",
            ciudad=up(g(r, idx.get("CIUDAD DESTINO"))) or "(SIN DATO)",
            ganancia=money(g(r, idx.get("GANANCIA"))), flete=money(g(r, idx.get("PRECIO FLETE"))),
            costo_dev=money(g(r, idx.get("COSTO DEVOLUCION FLETE"))),
            valor=money(g(r, idx.get("VALOR FACTURADO"))), novedad=up(g(r, idx.get("NOVEDAD")))))
for path, acc in prods_f:
    rows, idx = load_rows(path)
    fmin = min((parse_date(g(r, idx.get("FECHA"))) for r in rows if parse_date(g(r, idx.get("FECHA")))), default=None)
    bl = period_label(path, fmin)
    for r in rows:
        if es_prueba(g(r, idx.get("NOMBRE CLIENTE")), g(r, idx.get("TELÉFONO"))): continue
        est = g(r, idx.get("ESTATUS"))
        prod.append(dict(cuenta=acc, bimestre=bl, oid=str(g(r, idx.get("ID"))),
            fecha=parse_date(g(r, idx.get("FECHA"))), clase=classify(est),
            trans=up(g(r, idx.get("TRANSPORTADORA"))) or "(SIN DATO)",
            depto=up(g(r, idx.get("DEPARTAMENTO DESTINO"))) or "(SIN DATO)",
            ciudad=up(g(r, idx.get("CIUDAD DESTINO"))) or "(SIN DATO)",
            producto=str(g(r, idx.get("PRODUCTO")) or "(SIN DATO)").strip(),
            cantidad=money(g(r, idx.get("CANTIDAD"))) or 1,
            telefono=phone_key(g(r, idx.get("TELÉFONO"))),
            nombre=str(g(r, idx.get("NOMBRE CLIENTE")) or "").strip(),
            direccion=str(g(r, idx.get("DIRECCION")) or "").strip(),
            ganancia=money(g(r, idx.get("GANANCIA")))))

# ------------------------------------------------------------------ red de seguridad: duplicados
# Si el mismo export se descargó dos veces (o un mes aparece en dos archivos), la misma orden
# entraría doble y TODAS las cifras se inflarían en silencio. Se deduplica por ID de orden
# quedándose con la ÚLTIMA aparición (el archivo más reciente trae el estatus más actualizado)
# y se AVISA cuánto se unió, para que el usuario sepa que tenía archivos repetidos.
_antes = len(ped)
ped = list({(x["cuenta"], x["oid"]): x for x in ped}.values())
if len(ped) < _antes:
    print(f"⚠️ Duplicados detectados: {_antes - len(ped)} órdenes repetidas en los exports "
          f"'por pedido' (mismo ID en la misma cuenta). Las uní quedándome con la más reciente — "
          f"revisa si descargaste un archivo dos veces.")
_antes = len(prod)
# En 'por producto' una orden trae una fila por producto: el ID se repite legítimamente.
# El duplicado real es la MISMA fila (cuenta + orden + producto) otra vez.
prod = list({(x["cuenta"], x["oid"], x["producto"]): x for x in prod}.values())
if len(prod) < _antes:
    print(f"⚠️ Duplicados detectados: {_antes - len(prod)} filas repetidas en los exports "
          f"'por producto'. Las uní — revisa si descargaste un archivo dos veces.")

# enriquecimiento WhatsApp (opcional)
wp_label = defaultdict(set); wp_name = {}
if os.path.isdir(SRC):
    for fp in glob.glob(os.path.join(SRC, "*.csv")):
        base_fn = os.path.basename(fp).upper()
        if not (base_fn.startswith("ETIQUETA") or base_fn.startswith("CONTACTO")): continue
        with open(fp, encoding="utf-8", errors="replace") as f:
            for row in csv.DictReader(f):
                k = phone_key(row.get("phone"))
                if not k: continue
                lab = (row.get("labels") or "").strip()
                if lab:
                    for l in lab.split(","): wp_label[k].add(l.strip())
                nm = (row.get("saved_name") or row.get("public_name") or "").strip()
                if nm and k not in wp_name: wp_name[k] = nm

print(f"Cargado: {len(ped)} filas pedido | {len(prod)} filas producto | {len(wp_label)} tel con etiqueta WP")
cuentas = sorted(set(x["cuenta"] for x in ped) | set(x["cuenta"] for x in prod))

# ------------------------------------------------------------------ estilos
HDR = Font(bold=True, color="FFFFFF", size=10); HDRF = PatternFill("solid", fgColor="1F2A44")
TIT = Font(bold=True, size=14, color="1F2A44"); SUB = Font(bold=True, size=11, color="B8860B")
thin = Side(style="thin", color="DDDDDD"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c); cell.font = HDR; cell.fill = HDRF
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BORD
def pct(x): return f"{x*100:.1f}%"
def cop(v): return f"{CUR}" + f"{v:,.0f}".replace(",", ".")  # separador de miles con punto (Colombia)
def eff(counter):
    ent = counter["entregado"]; dev = counter["devolucion"]; act = ent + dev
    return ent, dev, act, (ent / act if act else 0.0)
def agg_status(records, keyfn):
    d = defaultdict(Counter); extra = defaultdict(lambda: defaultdict(float))
    for x in records:
        k = keyfn(x); d[k][x["clase"]] += 1
        for f in ("ganancia", "flete", "costo_dev", "valor"): extra[k][f] += x.get(f, 0)
    return d, extra

# ================================================================ LOGÍSTICA
wbL = Workbook(); wbL.remove(wbL.active)
ws = wbL.create_sheet("RESUMEN")
ws["B2"] = "MAESTRO LOGÍSTICA — DROPI"; ws["B2"].font = TIT
fechas = [x["fecha"] for x in ped if x["fecha"]]
ws["B3"] = (f"Periodo: {min(fechas):%b-%Y} a {max(fechas):%b-%Y}  |  Cuentas: {', '.join(cuentas)}"
            if fechas else "Periodo: s/d")
ws["B3"].font = Font(italic=True, color="666666")
scopes = [("GLOBAL", ped)] + [(f"Cuenta {c}", [x for x in ped if x["cuenta"] == c]) for c in cuentas] \
         if len(cuentas) > 1 else [("GLOBAL", ped)]
r = 5
for scope, recs in scopes:
    c = Counter(x["clase"] for x in recs); ent, dev, act, e = eff(c)
    gan = sum(x["ganancia"] for x in recs); cdev = sum(x["costo_dev"] for x in recs)
    ws.cell(r, 2, scope).font = SUB; r += 1
    for lab, val in [("Órdenes totales", len(recs)), ("Activas (entreg+devol)", act),
                     ("Entregadas", ent), ("Devoluciones", dev),
                     ("Canceladas/Rechazadas", c["cancelado"]), ("En tránsito", c["transito"]),
                     ("% ENTREGA", pct(e)), ("Ganancia neta acumulada", cop(gan)),
                     ("Costo devoluciones (flete)", cop(cdev))]:
        ws.cell(r, 2, lab); ws.cell(r, 4, val)
        if lab == "% ENTREGA":
            ws.cell(r, 2).font = Font(bold=True); ws.cell(r, 4).font = Font(bold=True, size=12, color="B8860B")
        r += 1
    r += 1
ws.column_dimensions["B"].width = 32; ws.column_dimensions["D"].width = 22

ws = wbL.create_sheet("EVOLUCIÓN")
ws["A1"] = "EVOLUCIÓN POR PERIODO (cronológico)"; ws["A1"].font = TIT
ws.append([]); hd = ["Periodo", "Cuenta", "Órdenes", "Activas", "Entregadas", "Devol.", "% Entrega", "Ganancia neta"]
ws.append(hd); style_header(ws, 3, len(hd))
by = defaultdict(list)
for x in ped: by[(x["bimestre"], x["cuenta"])].append(x)
for k in sorted(by):
    recs = by[k]; c = Counter(x["clase"] for x in recs); ent, dev, act, e = eff(c)
    ws.append([k[0], k[1], len(recs), act, ent, dev, pct(e), round(sum(x["ganancia"] for x in recs))])
for col, w in zip("ABCDEFGH", [12, 12, 10, 10, 11, 9, 11, 16]): ws.column_dimensions[col].width = w

dP, _ = agg_status(prod, lambda x: x["producto"])
unitsP = defaultdict(float)
for x in prod: unitsP[x["producto"]] += x["cantidad"]
ws = wbL.create_sheet("POR PRODUCTO")
ws["A1"] = "DESEMPEÑO POR PRODUCTO"; ws["A1"].font = TIT
ws.append([]); hd = ["Producto", "Líneas activas", "Entregadas", "Devol.", "% Entrega", "Unidades"]
ws.append(hd); style_header(ws, 3, len(hd))
for prd, c in sorted(dP.items(), key=lambda kv: -(kv[1]["entregado"] + kv[1]["devolucion"])):
    ent, dev, act, e = eff(c)
    if act == 0 and unitsP[prd] < 3: continue
    ws.append([prd[:55], act, ent, dev, pct(e), round(unitsP[prd])])
ws.column_dimensions["A"].width = 50
for col in "BCDEF": ws.column_dimensions[col].width = 13

dT, eT = agg_status(ped, lambda x: x["trans"])
ws = wbL.create_sheet("POR TRANSPORTADORA")
ws["A1"] = "DESEMPEÑO POR TRANSPORTADORA"; ws["A1"].font = TIT
ws.append([]); hd = ["Transportadora", "Activas", "Entregadas", "Devol.", "% Entrega", "% Devol.", "Costo devol."]
ws.append(hd); style_header(ws, 3, len(hd))
for t, c in sorted(dT.items(), key=lambda kv: -(kv[1]["entregado"] + kv[1]["devolucion"])):
    ent, dev, act, e = eff(c)
    if act < 5: continue
    ws.append([t, act, ent, dev, pct(e), pct(dev / act if act else 0), round(eT[t]["costo_dev"])])
ws.column_dimensions["A"].width = 22
for col in "BCDEFG": ws.column_dimensions[col].width = 13

dD, _ = agg_status(ped, lambda x: x["depto"])
ws = wbL.create_sheet("POR DEPARTAMENTO")
ws["A1"] = "DESEMPEÑO POR DEPARTAMENTO"; ws["A1"].font = TIT
ws.append([]); hd = ["Departamento", "Activas", "Entregadas", "Devol.", "% Entrega", "% Devol."]
ws.append(hd); style_header(ws, 3, len(hd))
for d, c in sorted(dD.items(), key=lambda kv: -(kv[1]["entregado"] + kv[1]["devolucion"])):
    ent, dev, act, e = eff(c)
    if act < 3: continue
    ws.append([d, act, ent, dev, pct(e), pct(dev / act if act else 0)])
ws.column_dimensions["A"].width = 24
for col in "BCDEF": ws.column_dimensions[col].width = 13

ct = defaultdict(Counter)
for x in ped: ct[(x["ciudad"], x["trans"])][x["clase"]] += 1
city_best = {}; city_tot = Counter()
for (city, tr), c in ct.items():
    ent, dev, act, e = eff(c); city_tot[city] += act
    if act >= 4 and (city not in city_best or e > city_best[city][2]
                     or (e == city_best[city][2] and act > city_best[city][1])):
        city_best[city] = (tr, act, e)
ws = wbL.create_sheet("MEJOR TRANSP x CIUDAD")
ws["A1"] = "MEJOR TRANSPORTADORA POR CIUDAD (mín. 4 órdenes activas)"; ws["A1"].font = TIT
ws.append([]); hd = ["Ciudad", "Total órdenes", "Mejor transportadora", "Órdenes", "% Entrega", "Recomendación"]
ws.append(hd); style_header(ws, 3, len(hd))
def reco(e):
    if e >= 0.85: return "EXCELENTE — enviar sin dudar"
    if e >= 0.70: return "BUENA — usar"
    if e >= 0.55: return "REGULAR — vigilar"
    return "MALA — evitar / solo anticipado"
for city in sorted(city_best, key=lambda c: -city_tot[c]):
    tr, act, e = city_best[city]
    ws.append([city, city_tot[city], tr, act, pct(e), reco(e)])
ws.column_dimensions["A"].width = 24; ws.column_dimensions["C"].width = 20; ws.column_dimensions["F"].width = 30
for col in "BDE": ws.column_dimensions[col].width = 13

nov = Counter(x["novedad"] for x in ped if x["novedad"] and x["novedad"] != "NONE")
totnov = sum(nov.values())
ws = wbL.create_sheet("NOVEDADES")
ws["A1"] = "PRINCIPALES NOVEDADES (causas de no-entrega)"; ws["A1"].font = TIT
ws.append([]); hd = ["Novedad", "Casos", "% del total"]; ws.append(hd); style_header(ws, 3, len(hd))
for n, c in nov.most_common(30):
    ws.append([n[:60], c, pct(c / totnov if totnov else 0)])
ws.column_dimensions["A"].width = 55; ws.column_dimensions["B"].width = 10; ws.column_dimensions["C"].width = 12

# ---- RESUMEN EJECUTIVO / RENTABILIDAD (primera hoja + insumo del PDF) ----
def suma(campo, filtro): return sum(x.get(campo, 0) for x in ped if filtro(x))
gan_realizada = suma("ganancia", lambda x: x["flujo"] == "realizado")
gan_en_camino = suma("ganancia", lambda x: x["flujo"] in ("en_camino", "en_camino_devol"))
if gan_en_camino <= 0:   # Dropi solo contabiliza ganancia al entregar; estimar en-camino con el promedio entregado
    _nreal = sum(1 for x in ped if x["flujo"] == "realizado")
    _avg = (gan_realizada / _nreal) if _nreal else 0
    gan_en_camino = _avg * sum(1 for x in ped if x["flujo"] in ("en_camino", "en_camino_devol"))
costo_dev     = suma("costo_dev", lambda x: x["flujo"] in ("devuelto", "en_camino_devol"))
util_dropi    = gan_realizada - costo_dev
n_flujo = Counter(x["flujo"] for x in ped)
gasto_pub = CFG.get("gasto_publicidad")
try: gasto_pub = float(gasto_pub) if gasto_pub not in (None, "") else None
except: gasto_pub = None
util_final = (util_dropi - gasto_pub) if gasto_pub is not None else None
if util_final is None:      veredicto, color = "FALTA GASTO DE PUBLICIDAD PARA EL VEREDICTO", "B8860B"
elif util_final > 0:        veredicto, color = "RENTABLE", "1E7A34"
else:                       veredicto, color = "NO RENTABLE", "B00020"

wsE = wbL.create_sheet("RESUMEN EJECUTIVO", 0)
wsE["B2"] = f"RESUMEN EJECUTIVO — {CFG.get('negocio','')} · DROPI"; wsE["B2"].font = Font(bold=True, size=16, color="1F2A44")
wsE["B3"] = (f"Periodo: {min(fechas):%d-%b-%Y} a {max(fechas):%d-%b-%Y}" if fechas else "Periodo: s/d")
wsE["B3"].font = Font(italic=True, color="666666")
def _row(ws, r, lab, val, bold=False, col="000000", size=11):
    ws.cell(r, 2, lab).font = Font(bold=bold, size=size)
    c = ws.cell(r, 4, val); c.font = Font(bold=bold, size=size, color=col)
    c.alignment = Alignment(horizontal="right")
r = 5
wsE.cell(r, 2, "ESTADO DE TUS ÓRDENES (dónde está la plata hoy)").font = SUB; r += 1
_row(wsE, r, "✅ Entregadas (cobradas)", n_flujo["realizado"]); r += 1
_row(wsE, r, "🚚 En camino (aún sin definir)", n_flujo["en_camino"]); r += 1
_row(wsE, r, "⚠️ En camino a devolución", n_flujo["en_camino_devol"]); r += 1
_row(wsE, r, "❌ Devueltas (perdidas)", n_flujo["devuelto"]); r += 1
_row(wsE, r, "🚫 Canceladas/Rechazadas (no cuentan)", n_flujo["cancelado"]); r += 2
wsE.cell(r, 2, "RENTABILIDAD").font = SUB; r += 1
_row(wsE, r, "Ganancia realizada (de lo entregado)", cop(gan_realizada), bold=True, col="1E7A34"); r += 1
_row(wsE, r, "(−) Costo de devoluciones (flete)", f"-{cop(costo_dev)}", col="B00020"); r += 1
_row(wsE, r, "= Utilidad Dropi (antes de publicidad)", cop(util_dropi), bold=True); r += 1
_row(wsE, r, "(−) Gasto de publicidad (Meta)",
     f"-{cop(gasto_pub)}" if gasto_pub is not None else "← FALTA (ponlo en _config_dropi.json o pásalo)",
     col="B00020"); r += 1
_row(wsE, r, "= UTILIDAD NETA FINAL",
     cop(util_final) if util_final is not None else "—", bold=True, size=13,
     col=("1E7A34" if (util_final or 0) > 0 else "B00020")); r += 1
_row(wsE, r, "Ganancia potencial en camino (estimada)", cop(gan_en_camino), col="666666"); r += 2
wsE.cell(r, 2, "VEREDICTO").font = Font(bold=True, size=12)
vc = wsE.cell(r, 4, veredicto); vc.font = Font(bold=True, size=13, color=color); vc.alignment = Alignment(horizontal="right")
wsE.column_dimensions["B"].width = 42; wsE.column_dimensions["D"].width = 34

pathL = os.path.join(OUT, "MAESTRO_LOGISTICA.xlsx"); wbL.save(pathL)
print("OK ->", pathL)

# ---- DOCUMENTO PDF: Resumen Ejecutivo de Rentabilidad ----
pdf_ok = False
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    pathPDF = os.path.join(OUT, "RESUMEN_EJECUTIVO.pdf")
    doc = SimpleDocTemplate(pathPDF, pagesize=letter, topMargin=1.6*cm, bottomMargin=1.4*cm,
                            leftMargin=1.8*cm, rightMargin=1.8*cm)
    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontSize=19, textColor=colors.HexColor("#1F2A44"), spaceAfter=2)
    Sb = ParagraphStyle("Sb", parent=ss["Normal"], fontSize=9.5, textColor=colors.HexColor("#666666"))
    Se = ParagraphStyle("Se", parent=ss["Heading2"], fontSize=12.5, textColor=colors.HexColor("#B8860B"), spaceBefore=12, spaceAfter=4)
    P = ParagraphStyle("P", parent=ss["Normal"], fontSize=10.3, leading=15)
    money_ = cop
    el = [Paragraph(f"Resumen Ejecutivo — {CFG.get('negocio','')} · Dropi", H),
          Paragraph((f"Periodo {min(fechas):%d-%b-%Y} a {max(fechas):%d-%b-%Y}  ·  "
                     f"Cuentas: {', '.join(cuentas)}  ·  Generado por golden-dropi-analisis") if fechas else "", Sb),
          Spacer(1, 8)]
    # Veredicto banner
    vb = Table([[veredicto]], colWidths=[17.4*cm])
    vb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#"+color)),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.white),("FONTSIZE",(0,0),(-1,-1),15),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("TOPPADDING",(0,0),(-1,-1),9),("BOTTOMPADDING",(0,0),(-1,-1),9)]))
    el += [vb, Spacer(1, 10)]
    tot = len(ped) or 1  # si solo hay export "por producto", ped=[]: evita ZeroDivision en los %
    est = [["Estado de las órdenes", "Cantidad", "% del total"],
           ["Entregadas (cobradas)", n_flujo["realizado"], f"{n_flujo['realizado']/tot*100:.1f}%"],
           ["En camino (sin definir)", n_flujo["en_camino"], f"{n_flujo['en_camino']/tot*100:.1f}%"],
           ["En camino a devolución", n_flujo["en_camino_devol"], f"{n_flujo['en_camino_devol']/tot*100:.1f}%"],
           ["Devueltas (perdidas)", n_flujo["devuelto"], f"{n_flujo['devuelto']/tot*100:.1f}%"],
           ["Canceladas/Rechazadas", n_flujo["cancelado"], f"{n_flujo['cancelado']/tot*100:.1f}%"]]
    t1 = Table(est, colWidths=[9*cm, 4.2*cm, 4.2*cm])
    t1.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F2A44")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9.5),("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F5F5")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#DDDDDD")),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    el += [Paragraph("Dónde está tu plata hoy", Se), t1]
    pl = [["Rentabilidad", ""],
          ["Ganancia realizada (entregado)", money_(gan_realizada)],
          ["(−) Costo de devoluciones (flete)", "-"+money_(costo_dev)],
          ["= Utilidad Dropi (antes de publicidad)", money_(util_dropi)],
          ["(−) Gasto de publicidad (Meta)", ("-"+money_(gasto_pub)) if gasto_pub is not None else "FALTA"],
          ["= UTILIDAD NETA FINAL", money_(util_final) if util_final is not None else "—"],
          ["Ganancia potencial en camino (est.)", money_(gan_en_camino)]]
    t2 = Table(pl, colWidths=[12*cm, 5.4*cm])
    t2.setStyle(TableStyle([("SPAN",(0,0),(1,0)),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#B8860B")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,3),(-1,3),"Helvetica-Bold"),("FONTNAME",(0,5),(-1,5),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),10),("FONTSIZE",(0,5),(-1,5),12),("ALIGN",(1,0),(1,-1),"RIGHT"),
        ("TEXTCOLOR",(1,1),(1,1),colors.HexColor("#1E7A34")),("TEXTCOLOR",(1,2),(1,2),colors.HexColor("#B00020")),
        ("TEXTCOLOR",(1,4),(1,4),colors.HexColor("#B00020")),
        ("TEXTCOLOR",(1,5),(1,5),colors.HexColor("#1E7A34") if (util_final or 0)>0 else colors.HexColor("#B00020")),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#DDDDDD")),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    el += [Paragraph("Rentabilidad", Se), t2]
    if gasto_pub is None:
        el += [Spacer(1,6), Paragraph("<b>Para cerrar el veredicto falta tu gasto de publicidad</b> del periodo "
            "(Meta). Ponlo en <font face='Courier'>_config_dropi.json</font> (clave "
            "<font face='Courier'>gasto_publicidad</font>) o pásalo, y el neto y el veredicto se calculan solos. "
            "Tu dashboard de gasto ya lo tiene: ahí se puede conectar.", P)]
    cg = Counter(x["clase"] for x in ped); ent, dev, act, ef = eff(cg)
    el += [Paragraph("Contexto de entregas", Se),
           Paragraph(f"Efectividad de entrega: <b>{ef*100:.1f}%</b> ({ent:,} entregadas de {act:,} cerradas). "
                     f"El detalle por producto, transportadora, ciudad y las novedades está en "
                     f"<b>MAESTRO_LOGISTICA.xlsx</b>; la base de clientes en <b>MAESTRO_CONTACTOS.xlsx</b>.", P)]
    doc.build(el)
    print("OK ->", pathPDF); pdf_ok = True
except ImportError:
    print("Aviso: sin 'reportlab' no se generó el PDF (pip install reportlab). El RESUMEN EJECUTIVO está en el Excel.")

# ================================================================ CONTACTOS
cli = defaultdict(lambda: dict(nombre="", depto="", ciudad="", direccion="", prods=set(),
                               cuentas=set(), ordenes={}, fmin=None, fmax=None))
for x in prod:
    k = x["telefono"]
    if not k or len(k) < 7: continue
    c = cli[k]
    if x["nombre"] and (not c["nombre"] or (x["fecha"] and (c["fmax"] is None or x["fecha"] >= c["fmax"]))):
        c["nombre"] = x["nombre"]
    if x["ciudad"] != "(SIN DATO)": c["ciudad"] = x["ciudad"]; c["depto"] = x["depto"]
    if x["direccion"]: c["direccion"] = x["direccion"]
    if x["producto"] != "(SIN DATO)": c["prods"].add(x["producto"][:30])
    c["cuentas"].add(x["cuenta"])
    if x["oid"] not in c["ordenes"] or x["clase"] == "entregado": c["ordenes"][x["oid"]] = x["clase"]
    if x["fecha"]:
        c["fmin"] = x["fecha"] if c["fmin"] is None else min(c["fmin"], x["fecha"])
        c["fmax"] = x["fecha"] if c["fmax"] is None else max(c["fmax"], x["fecha"])
def segmento(np, e, resueltos):
    # Sin órdenes cerradas (todas en tránsito o canceladas) la efectividad es indefinida:
    # NEUTRO, no RIESGO. Antes marcaba "mucha devolución" a quien no tenía ni una devolución.
    if resueltos == 0: return "NEUTRO"
    if np >= 3 and e >= 0.7: return "VIP (recurrente confiable)"
    if e >= 0.7: return "BUENO"
    if e < 0.5: return "RIESGO (mucha devolución)"
    return "NEUTRO"

wbC = Workbook(); wbC.remove(wbC.active)
ws = wbC.create_sheet("CLIENTES")
ws["A1"] = "MAESTRO DE CONTACTOS — CLIENTES CON PEDIDOS (sin duplicados)"; ws["A1"].font = TIT
ws.append([]); hd = ["Teléfono", "Nombre", "Depto", "Ciudad", "Productos", "Pedidos", "Entregados",
      "Devol.", "% Efectividad", "Segmento", "Etiqueta WhatsApp", "Última compra", "Cuenta"]
ws.append(hd); style_header(ws, 3, len(hd))
rowsC = []
for k, c in cli.items():
    ents = sum(1 for v in c["ordenes"].values() if v == "entregado")
    devs = sum(1 for v in c["ordenes"].values() if v == "devolucion")
    npd = len(c["ordenes"]); den = ents + devs; e = (ents / den) if den else 0
    rowsC.append((k, c, npd, ents, devs, e))
rowsC.sort(key=lambda t: (-t[2], -t[5]))
for k, c, npd, ents, devs, e in rowsC:
    nombre = c["nombre"] or wp_name.get(k, "")   # fallback: nombre guardado en WhatsApp
    ws.append([k, nombre[:30], c["depto"], c["ciudad"], ", ".join(sorted(c["prods"]))[:60],
        npd, ents, devs, pct(e) if (ents + devs) else "-", segmento(npd, e, ents + devs),
        ", ".join(sorted(wp_label.get(k, [])))[:40], f"{c['fmax']:%Y-%m-%d}" if c["fmax"] else "",
        "+".join(sorted(c["cuentas"]))])
for col, w in zip("ABCDEFGHIJKLM", [13, 26, 16, 18, 40, 8, 10, 8, 12, 26, 26, 13, 12]):
    ws.column_dimensions[col].width = w

# leads del bot (opcional)
ws = wbC.create_sheet("LEADS NO CLIENTES (bot)")
ws["A1"] = "LEADS DEL BOT QUE NO COMPRARON"; ws["A1"].font = TIT
ws.append([]); hd = ["Teléfono", "Nombre", "Ciudad", "Producto interés", "Últi. interacción", "Etiqueta WP"]
ws.append(hd); style_header(ws, 3, len(hd)); nlead = 0
leadfp = None
if os.path.isdir(SRC):
    for fp in glob.glob(os.path.join(SRC, "*.csv")):
        if os.path.basename(fp).upper().startswith("NO CLIENTE"): leadfp = fp; break
if leadfp:
    with open(leadfp, encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            k = phone_key(row.get("phone"))
            nm = (row.get("name") or row.get("first_name") or "").strip()
            city = (row.get("city") or row.get("Ciudad") or "").strip()
            inter = (row.get("Producto a activar") or row.get("interest") or row.get("Producto Remarketing") or "").strip()
            last = (row.get("last_interaction") or "").strip()[:10]
            ws.append([k, nm[:26], city[:20], inter[:30], last, ", ".join(sorted(wp_label.get(k, [])))[:35]]); nlead += 1
for col, w in zip("ABCDEF", [13, 26, 20, 30, 13, 35]): ws.column_dimensions[col].width = w

# posibles / mensaje masivo (opcional, de un BASE*.xlsx en _FUENTES con hoja "POSIBLE*")
nposibles = 0
if os.path.isdir(SRC):
    for bmp in glob.glob(os.path.join(SRC, "*.xlsx")):
        try: wbb = openpyxl.load_workbook(bmp, read_only=True, data_only=True)
        except Exception: continue
        sh = next((s for s in wbb.sheetnames if "POSIBLE" in s.upper()), None)
        if not sh: wbb.close(); continue
        wsp = wbC.create_sheet("POSIBLES (msj masivo)")
        wsp["A1"] = "POSIBLES / LEADS PARA MENSAJE MASIVO"; wsp["A1"].font = TIT
        rws = [r for r in wbb[sh].iter_rows(values_only=True) if any(x is not None for x in r)]
        hdr = [str(x).strip() if x else "" for x in rws[0]]
        wsp.append([]); wsp.append(hdr); style_header(wsp, 3, len(hdr)); seen = set()
        for r in rws[1:]:
            k = phone_key(g(r, 0))
            if k in seen or es_prueba(g(r, 1), g(r, 0)): continue
            seen.add(k); wsp.append([str(x).strip() if x is not None else "" for x in r]); nposibles += 1
        for col, w in zip("ABCDEFGHIJKL", [15, 24, 16, 16, 26, 10, 10, 10, 10, 10, 14, 12]):
            wsp.column_dimensions[col].width = w
        wbb.close(); break

ws = wbC.create_sheet("RESUMEN", 0)
ws["B2"] = "MAESTRO DE CONTACTOS — RESUMEN"; ws["B2"].font = TIT
segc = Counter(segmento(npd, e, ents + devs) for _, _, npd, ents, devs, e in rowsC)
r = 4
for lab, v in [("Clientes únicos (con pedido)", len(rowsC)),
               ("  · VIP recurrentes", segc["VIP (recurrente confiable)"]),
               ("  · Buenos", segc["BUENO"]), ("  · En riesgo (devolución)", segc["RIESGO (mucha devolución)"]),
               ("  · Neutros", segc["NEUTRO"]),
               ("Clientes con etiqueta WhatsApp", sum(1 for k, *_ in rowsC if k in wp_label)),
               ("Leads del bot sin compra", nlead), ("Posibles para mensaje masivo", nposibles)]:
    ws.cell(r, 2, lab); ws.cell(r, 4, v); r += 1
ws.column_dimensions["B"].width = 34; ws.column_dimensions["D"].width = 12

pathC = os.path.join(OUT, "MAESTRO_CONTACTOS.xlsx"); wbC.save(pathC)
print("OK ->", pathC)
print(f"Clientes únicos: {len(rowsC)} | Leads bot: {nlead} | Posibles: {nposibles}")
