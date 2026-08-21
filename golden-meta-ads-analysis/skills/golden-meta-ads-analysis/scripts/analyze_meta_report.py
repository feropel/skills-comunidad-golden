"""
analyze_meta_report.py — Lee y analiza un archivo Excel (.xlsx/.xls) o CSV de Meta Ads.

Detecta automáticamente el formato (Excel: Creative Reporting, Raw Data, Formatted;
CSV: header en distintas filas), identifica columnas, filtra solo campañas de venta,
y automatiza 4 de las 8 capas del análisis (ver SKILL.md Paso 4):
  - Capa 1: tipos de campaña      → analyze_by_campaign_type()
  - Capa 2: creativos (anuncios)  → analyze_by_ad()
  - Capa 3: demografía (edad×sexo) → analyze_by_demographics()
  - Capa 4: plataforma y ubicación → analyze_by_placement()

Las capas 5-8 (landings, tipo/formato de creativo, embudo de conversión, tendencia
temporal) dependen de columnas muy variables entre exports (URL de destino, fechas)
y se calculan inline sobre el mismo DataFrame en cada análisis — no tienen función
dedicada aquí porque su forma cambia según qué desglose trajo el archivo.

Uso:
    from analyze_meta_report import load_meta_report, analyze_all_layers

    df = load_meta_report('/path/to/file.xlsx')   # o '/path/to/file.csv'
    analysis = analyze_all_layers(df, ticket=None)
"""

import pandas as pd
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')


def load_meta_report(filepath):
    """
    Carga un archivo Excel O CSV de Meta Ads detectando automáticamente:
    - Formato de archivo (.csv vs .xlsx/.xls)
    - Hoja correcta (solo Excel)
    - Fila del header
    - Tipo de formato

    Returns: DataFrame procesado y limpio
    """
    filepath = Path(filepath)

    if filepath.suffix.lower() == ".csv":
        return _load_csv_report(filepath)
    return _load_excel_report(filepath)


def _load_csv_report(filepath):
    """Carga un CSV exportado de Ads Manager, probando encoding y fila de header."""
    df = None
    last_err = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        for header_row in [0, 1, 2, 3, 8]:
            try:
                test_df = pd.read_csv(filepath, header=header_row, nrows=5, encoding=encoding)
                named_cols = [c for c in test_df.columns
                              if not str(c).startswith('Unnamed') and
                              not str(c).replace('.', '').replace('-', '').isdigit()]
                if len(named_cols) >= 10:
                    df = pd.read_csv(filepath, header=header_row, encoding=encoding)
                    break
            except Exception as e:
                last_err = e
                continue
        if df is not None:
            break

    if df is None:
        # Último recurso: header en fila 0 con utf-8 (o lanza el error real)
        try:
            df = pd.read_csv(filepath, header=0, encoding="utf-8-sig")
        except Exception:
            raise ValueError(
                f"No se pudo leer el CSV '{filepath.name}'. Verifica que sea un export "
                f"válido de Ads Manager (delimitado por comas). Error original: {last_err}"
            )

    return _apply_hierarchy_ffill(df)


def _load_excel_report(filepath):
    from openpyxl import load_workbook

    # Detectar hojas
    wb = load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    
    # Prioridad de hojas
    target_sheet = None
    for priority in ['Creative Reporting', 'Raw Data Report', 'Formatted Report']:
        if priority in sheets:
            target_sheet = priority
            break
    if target_sheet is None:
        target_sheet = sheets[0]
    
    # Detectar fila del header probando varias
    df = None
    for header_row in [0, 1, 2, 3, 8]:
        try:
            test_df = pd.read_excel(filepath, sheet_name=target_sheet, header=header_row, nrows=5)
            named_cols = [c for c in test_df.columns 
                         if not str(c).startswith('Unnamed') and 
                         not str(c).replace('.','').replace('-','').isdigit()]
            if len(named_cols) >= 10:
                df = pd.read_excel(filepath, sheet_name=target_sheet, header=header_row)
                break
        except:
            continue
    
    if df is None:
        df = pd.read_excel(filepath, sheet_name=target_sheet, header=0)

    return _apply_hierarchy_ffill(df)


def _apply_hierarchy_ffill(df):
    """Forward-fill jerarquía (campaña/adset) si el reporte trae NaN en niveles superiores.
    Compartido entre carga de Excel y de CSV — mismo formato de columnas en ambos."""
    if 'Nombre de la campaña' in df.columns:
        df['Campaña_ff'] = df['Nombre de la campaña'].ffill()
    if 'Nombre del conjunto de anuncios' in df.columns:
        df['Adset_ff'] = df['Nombre del conjunto de anuncios'].ffill()
    return df


def detect_currency(df):
    """Detecta la moneda del gasto."""
    for col in df.columns:
        if 'Importe gastado' in str(col) or 'Amount spent' in str(col):
            if 'USD' in str(col):
                return 'USD', col
            elif 'COP' in str(col):
                return 'COP', col
            else:
                return 'unknown', col
    return None, None


def filter_sales_campaigns(df):
    """Filtra solo campañas de tipo Compras en el sitio web."""
    if 'Tipo de resultado' not in df.columns:
        return df  # Si no hay columna, asumir todo es venta
    
    # Campañas que tienen alguna fila con tipo "Compras en el sitio web"
    sales_mask = df['Tipo de resultado'] == 'Compras en el sitio web'
    sales_campaigns = df[sales_mask]['Nombre de la campaña'].dropna().unique()
    
    if 'Campaña_ff' in df.columns:
        return df[df['Campaña_ff'].isin(sales_campaigns)].copy()
    return df[df['Nombre de la campaña'].isin(sales_campaigns)].copy()


def classify_campaign_type(name):
    """Clasifica una campaña por su nombre."""
    if pd.isna(name):
        return 'OTRO'
    n = str(name).upper()
    
    if 'WPP' in n or 'WHATSAPP' in n or 'CHATEA' in n or 'CHATBOT' in n:
        return 'TRAFICO_WPP'
    if 'TRAFICO' in n or 'TRÁFICO' in n or 'TRAFFIC' in n:
        return 'TRAFICO_WEB'
    if 'RETARGETING' in n or 'REMARKETING' in n or 'RT/' in n or '/RT' in n:
        return 'RETARGETING'
    if 'ADVANTAGE' in n or 'AD+' in n or 'ASC' in n:
        return 'ADVANTAGE_PLUS'
    if 'CATALOG' in n or 'CATÁLOGO' in n or 'DPA' in n:
        return 'CATALOG_DPA'
    if 'BACK UP' in n or 'BACKUP' in n:
        return 'BACKUP'
    if 'OPEN' in n or 'VENTA' in n or 'WEB' in n:
        return 'VENTA_OPEN'
    return 'OTRO'


def analyze_by_campaign_type(df, gasto_col, compras_col, ticket=None):
    """Capa 1: Performance por tipo de campaña."""
    df = df.copy()
    campaign_col = 'Campaña_ff' if 'Campaña_ff' in df.columns else 'Nombre de la campaña'
    df['_tipo'] = df[campaign_col].apply(classify_campaign_type)
    
    agg = df.groupby('_tipo').agg(
        gasto=(gasto_col, 'sum'),
        compras=(compras_col, 'sum')
    ).reset_index()
    agg['CPA'] = (agg['gasto'] / agg['compras']).round(0)
    if ticket:
        agg['ROAS'] = (agg['compras'] * ticket / agg['gasto']).round(2)
    agg['pct_compras'] = (agg['compras'] / agg['compras'].sum() * 100).round(1)
    
    return agg.sort_values('gasto', ascending=False)


def analyze_by_ad(df, gasto_col, compras_col, ticket=None, min_gasto=50000):
    """Capa 2: Performance por anuncio (creativo)."""
    df = df.copy()
    
    # Filtrar solo nivel ad si existe
    if 'Nivel de la entrega' in df.columns:
        df = df[df['Nivel de la entrega'] == 'ad']
    
    agg = df.groupby('Nombre del anuncio').agg(
        gasto=(gasto_col, 'sum'),
        compras=(compras_col, 'sum'),
        ad_ids=('Identificador del anuncio', 'nunique')
    ).reset_index()
    
    agg['CPA'] = (agg['gasto'] / agg['compras']).round(0)
    if ticket:
        agg['ROAS'] = (agg['compras'] * ticket / agg['gasto']).round(2)
    
    # Métricas adicionales si existen
    for col_name, agg_col in [
        ('CTR único (porcentaje de clics en el enlace)', 'CTR'),
        ('CPC (costo por clic en el enlace)', 'CPC'),
        ('CPM (costo por mil impresiones)', 'CPM'),
        ('Frecuencia', 'Frecuencia')
    ]:
        if col_name in df.columns:
            # Ponderado por gasto
            weighted = df.groupby('Nombre del anuncio').apply(
                lambda g: (g[col_name] * g[gasto_col]).sum() / g[gasto_col].sum() if g[gasto_col].sum() > 0 else 0
            )
            agg[agg_col] = agg['Nombre del anuncio'].map(weighted).round(2)
    
    # Hold rate y view through si hay datos de video
    if 'Reproducciones de video de 3 segundos' in df.columns and 'Reproducciones de video hasta el 50%' in df.columns:
        video_agg = df.groupby('Nombre del anuncio').agg(
            v3s=('Reproducciones de video de 3 segundos', 'sum'),
            v50=('Reproducciones de video hasta el 50%', 'sum')
        )
        if 'Reproducciones de video hasta el 100%' in df.columns:
            video_agg['v100'] = df.groupby('Nombre del anuncio')['Reproducciones de video hasta el 100%'].sum()
            agg['view_through_100'] = (agg['Nombre del anuncio'].map(video_agg['v100'] / video_agg['v3s'] * 100)).round(1)
        agg['hold_rate_50'] = (agg['Nombre del anuncio'].map(video_agg['v50'] / video_agg['v3s'] * 100)).round(1)
    
    return agg[agg['gasto'] >= min_gasto].sort_values('gasto', ascending=False)


def analyze_by_demographics(df, gasto_col, compras_col, ticket=None):
    """Capa 3: Performance por edad y sexo."""
    if 'Edad' not in df.columns:
        return None
    
    # Filtrar filas con desglose real (no totales 'All')
    df_d = df[(df['Edad'] != 'All') & (df['Edad'] != 'Unknown') & df['Edad'].notna()].copy()
    
    # Por sexo
    sexo_agg = None
    if 'Sexo' in df.columns:
        sexo_agg = df_d.groupby('Sexo').agg(
            gasto=(gasto_col, 'sum'),
            compras=(compras_col, 'sum')
        ).reset_index()
        sexo_agg['CPA'] = (sexo_agg['gasto'] / sexo_agg['compras']).round(0)
        if ticket:
            sexo_agg['ROAS'] = (sexo_agg['compras'] * ticket / sexo_agg['gasto']).round(2)
        sexo_agg['pct_compras'] = (sexo_agg['compras'] / sexo_agg['compras'].sum() * 100).round(1)
    
    # Por edad
    edad_agg = df_d.groupby('Edad').agg(
        gasto=(gasto_col, 'sum'),
        compras=(compras_col, 'sum')
    ).reset_index()
    edad_agg['CPA'] = (edad_agg['gasto'] / edad_agg['compras']).round(0)
    if ticket:
        edad_agg['ROAS'] = (edad_agg['compras'] * ticket / edad_agg['gasto']).round(2)
    edad_agg['pct_compras'] = (edad_agg['compras'] / edad_agg['compras'].sum() * 100).round(1)
    
    # Combo edad x sexo
    combo_agg = None
    if 'Sexo' in df.columns:
        combo_agg = df_d.groupby(['Edad', 'Sexo']).agg(
            gasto=(gasto_col, 'sum'),
            compras=(compras_col, 'sum')
        ).reset_index()
        combo_agg = combo_agg[combo_agg['compras'] >= 5]  # filtro de relevancia
        combo_agg['CPA'] = (combo_agg['gasto'] / combo_agg['compras']).round(0)
        if ticket:
            combo_agg['ROAS'] = (combo_agg['compras'] * ticket / combo_agg['gasto']).round(2)
    
    return {
        'por_sexo': sexo_agg,
        'por_edad': edad_agg.sort_values('compras', ascending=False),
        'por_edad_x_sexo': combo_agg.sort_values('compras', ascending=False) if combo_agg is not None else None
    }


def analyze_by_placement(df, gasto_col, compras_col, ticket=None):
    """Capa 4: Performance por plataforma y ubicación."""
    if 'Plataforma' not in df.columns and 'Ubicación' not in df.columns:
        return None
    
    result = {}
    
    if 'Plataforma' in df.columns:
        plat = df.groupby('Plataforma').agg(
            gasto=(gasto_col, 'sum'),
            compras=(compras_col, 'sum')
        ).reset_index()
        plat['CPA'] = (plat['gasto'] / plat['compras']).round(0)
        if ticket:
            plat['ROAS'] = (plat['compras'] * ticket / plat['gasto']).round(2)
        plat['pct_compras'] = (plat['compras'] / plat['compras'].sum() * 100).round(1)
        result['por_plataforma'] = plat.sort_values('gasto', ascending=False)
    
    if 'Ubicación' in df.columns:
        ubi = df.groupby('Ubicación').agg(
            gasto=(gasto_col, 'sum'),
            compras=(compras_col, 'sum')
        ).reset_index()
        ubi['CPA'] = (ubi['gasto'] / ubi['compras']).round(0)
        if ticket:
            ubi['ROAS'] = (ubi['compras'] * ticket / ubi['gasto']).round(2)
        ubi['pct_compras'] = (ubi['compras'] / ubi['compras'].sum() * 100).round(1)
        result['por_ubicacion'] = ubi[ubi['compras'] >= 5].sort_values('compras', ascending=False)
    
    return result


def analyze_all_layers(df, ticket=None, min_gasto_ad=50000):
    """
    Ejecuta el análisis en 6 capas y devuelve un dict con todos los resultados.
    """
    # Detectar columnas
    moneda, gasto_col = detect_currency(df)
    if gasto_col is None:
        raise ValueError("No se encontró columna de gasto en el archivo")
    
    # Columna de compras: priorizar 'Compras', luego 'Resultados'
    compras_col = None
    for col in ['Compras', 'Resultados', 'Compras en el sitio web']:
        if col in df.columns:
            compras_col = col
            break
    if compras_col is None:
        raise ValueError("No se encontró columna de compras en el archivo")
    
    # Filtrar solo campañas de venta
    sales_df = filter_sales_campaigns(df)
    
    # Totales globales
    gasto_total = sales_df[gasto_col].sum()
    compras_total = sales_df[compras_col].sum()
    cpa_promedio = gasto_total / compras_total if compras_total > 0 else 0
    
    # Periodo
    periodo = None
    if 'Inicio del informe' in sales_df.columns and 'Fin del informe' in sales_df.columns:
        inicio = sales_df['Inicio del informe'].dropna().min()
        fin = sales_df['Fin del informe'].dropna().max()
        periodo = f"{inicio} a {fin}"
    
    return {
        'metadata': {
            'moneda': moneda,
            'periodo': periodo,
            'campañas_unicas': sales_df['Nombre de la campaña'].nunique() if 'Nombre de la campaña' in sales_df.columns else 0,
            'adsets_unicos': sales_df['Nombre del conjunto de anuncios'].nunique() if 'Nombre del conjunto de anuncios' in sales_df.columns else 0,
            'ads_unicos': sales_df['Nombre del anuncio'].nunique() if 'Nombre del anuncio' in sales_df.columns else 0,
        },
        'totales': {
            'gasto_total': round(gasto_total, 0),
            'compras_total': round(compras_total, 0),
            'cpa_promedio': round(cpa_promedio, 0)
        },
        'capa_1_tipos_campaña': analyze_by_campaign_type(sales_df, gasto_col, compras_col, ticket),
        'capa_2_creativos': analyze_by_ad(sales_df, gasto_col, compras_col, ticket, min_gasto_ad),
        'capa_3_demografia': analyze_by_demographics(sales_df, gasto_col, compras_col, ticket),
        'capa_4_placement': analyze_by_placement(sales_df, gasto_col, compras_col, ticket),
    }


def print_analysis_summary(analysis):
    """Imprime resumen del análisis."""
    print("="*80)
    print("ANÁLISIS DE META ADS")
    print("="*80)
    
    m = analysis['metadata']
    t = analysis['totales']
    
    print(f"\nPeriodo: {m['periodo']}")
    print(f"Moneda: {m['moneda']}")
    print(f"Campañas: {m['campañas_unicas']} | Adsets: {m['adsets_unicos']} | Ads: {m['ads_unicos']}")
    print(f"\nTotales:")
    print(f"  Gasto: ${t['gasto_total']:,.0f}")
    print(f"  Compras: {t['compras_total']:,.0f}")
    print(f"  CPA promedio: ${t['cpa_promedio']:,.0f}")
    
    print(f"\n--- CAPA 1: Por tipo de campaña ---")
    print(analysis['capa_1_tipos_campaña'].to_string())
    
    print(f"\n--- CAPA 2: Top creativos ---")
    print(analysis['capa_2_creativos'].head(15).to_string())
    
    if analysis['capa_3_demografia']:
        print(f"\n--- CAPA 3: Demografía por edad ---")
        print(analysis['capa_3_demografia']['por_edad'].to_string())
        if analysis['capa_3_demografia']['por_sexo'] is not None:
            print(f"\n--- Por sexo ---")
            print(analysis['capa_3_demografia']['por_sexo'].to_string())
    
    if analysis['capa_4_placement']:
        if 'por_plataforma' in analysis['capa_4_placement']:
            print(f"\n--- CAPA 4: Por plataforma ---")
            print(analysis['capa_4_placement']['por_plataforma'].to_string())
